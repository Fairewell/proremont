import openpyxl
import json


def export_sheet_name(name):
    wb = openpyxl.load_workbook(name + ".xlsx")

    # Initialize a list to store the results
    results = []

    # Iterate through each sheet
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        prices = []

        # Find the column index for the header "Цены (руб.)"
        for col in range(1, ws.max_column + 1):
            if ws.cell(row=1, column=col).value == "Цены (руб.)":
                prices = [
                    ws.cell(row=row, column=col).value
                    for row in range(2, ws.max_row + 1)
                    if ws.cell(row=row, column=col).value
                ]

        # Calculate the average price
        average_price = sum(prices) / len(prices) if prices else 0
        print(f"SHEETNAME: {sheet}\nAVER.PRICE: {average_price}")

        # Append the results to the list
        results.append(
            {
                "id": wb.sheetnames.index(sheet) + 1,
                "name": sheet,
                "price": average_price,
            }
        )

    # Save the results in a JSON file
    with open("average_prices.json", "w", encoding="UTF-8") as json_file:
        json.dump(results, json_file, ensure_ascii=False, indent=4)


name = "Прайс"
# export_sheet_name(name)

title = "Сантехнические работы"

# Table headers
headers = ["Наименование работ", "Единица изм.", "Цены (руб.)"]

# Example text
text = """
Разводка водопровода	точка	1750
Монтаж + демонтаж радиатора (бимет.)	шт.	4000
Замена стояка (канализ.)	этаж	6000
Замена стояка (водопровод.)	этаж	3500
Замена стояка (отопление.)	этаж	3500
Устройство временного водоснабжения	комп.	2000
Устройство сантехнической штробы	п.м.	750
Заделка сантехнической штробы	п.м.	200
Установка шарового крана	шт.	550
Установка фильтра грубой очистки	шт.	550
Установка счетчика воды	шт.	2000
Переделка узла подводки полотенцесушителя	комп.	10000
Установка фильтра тонкой очистки с регулятором давления	шт.	1750
Установка коллектора (комплекс работ)	шт.	2200
Прокладка труб ХГВ (м/пласт, п/пропилен, п/этилен)	п.м.	350
Прокладка канализационных труб (ПВХ)	п.м.	450
Врезка в канализационный стояк	шт.	2700
Врезка в водопроводный стояк	шт.	5500
Теплоизоляция труб	п.м.	100
Установка накопительного водонагревателя (бойлера)	шт.	4500
Установка проточного водонагревателя	шт.	3500
Установка инсталляции	шт.	3000
Установка подвесного унитаза	шт.	5500
Установка биде	шт.	2500
Установка унитаза	шт.	2000
Установка писсуара	шт.	3500
Установка умывальника	шт.	3000
Установка "мойдодыра"	шт.	3500
Установка смесителя	шт.	1200
Установка гигиенического душа	шт.	1200
Установка полотенцесушителя	шт.	4000
Установка стиральной машины	шт.	2500
Установка ванны	шт.	4500
Установка смесителя ванны на штанге	шт.	2500
Установка экрана ванны	шт.	1000
Установка гидромассажной ванны	шт.	7500
Установка душевой кабины с поддоном	шт.	15000
Установка душевой панели	шт.	2000
Установка поддона душевой кабины	комп.	2500
Установка дверок душевой кабины	шт.	4500
Установка трапа	шт.	1200
Установка шкафа	шт.	3500
Установка системы контроля протечки воды (ХГВ)	комп.	8500
Установка аксессуаров	шт.	450
"""


def text_to_excel(text, title, headers):
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = title
    sheet.append(headers)
    # Remove empty lines
    lines = [line for line in text.strip().split("\n") if line.strip()]

    # Add data to the table
    for i in range(0, len(lines), 3):
        try:
            row = [lines[i], lines[i + 1], int(lines[i + 2])]
            sheet.append(row)
        except (ValueError, IndexError):
            pass

    # Save the Excel workbook
    wb.save("./excel таблицы/" + sheet.title + ".xlsx")


text_to_excel(text, title, headers)
